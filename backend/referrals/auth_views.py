"""
Authentication views for the referrals application.
Includes rate limiting, structured logging, and optimized queries.
"""
import logging
import re
from io import BytesIO
from collections import defaultdict
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.utils import timezone
from django.http import HttpResponse
from django_ratelimit.decorators import ratelimit
from django.utils.decorators import method_decorator

from .models import Sympathizer, LevelLabel, Department, Municipality
from .services.email import EmailService

logger = logging.getLogger(__name__)


class CheckUserView(APIView):
    """Check if a user exists by cedula."""

    @method_decorator(ratelimit(key='ip', rate='20/m', method='POST', block=True))
    def post(self, request):
        cedula = request.data.get('cedula')
        if not cedula:
            return Response({'error': 'La cedula es requerida'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sympathizer = Sympathizer.objects.select_related('user').get(cedula=cedula)

            # Determine if the user has a valid account setup
            has_user = False
            if sympathizer.user:
                if sympathizer.user.is_active and sympathizer.user.has_usable_password():
                    has_user = True

            return Response({
                'exists': True,
                'has_user': has_user,
                'email': sympathizer.email,
                'masked_email': self._mask_email(sympathizer.email) if sympathizer.email else None
            })
        except Sympathizer.DoesNotExist:
            return Response({'exists': False}, status=status.HTTP_404_NOT_FOUND)

    @staticmethod
    def _mask_email(email: str) -> str:
        """Mask email for privacy."""
        if not email:
            return ""
        try:
            user_part, domain = email.split('@')
            masked = f"{user_part[:2]}{'*' * min(len(user_part) - 2, 5)}@{domain}"
            return masked
        except ValueError:
            return email


class RequestPasswordSetupView(APIView):
    """Request password setup email."""

    @method_decorator(ratelimit(key='ip', rate='5/m', method='POST', block=True))
    def post(self, request):
        cedula = request.data.get('cedula')
        if not cedula:
            return Response({'error': 'La cedula es requerida'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sympathizer = Sympathizer.objects.select_related('user').get(cedula=cedula)

            if not sympathizer.email:
                logger.warning(f"Password setup requested for user without email: {cedula[:4]}***")
                return Response(
                    {'error': 'No hay correo electronico registrado para este usuario'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Create an inactive user if not exists
            user = sympathizer.user
            if not user:
                user = User.objects.create_user(
                    username=cedula,
                    email=sympathizer.email,
                    is_active=False
                )
                user.set_unusable_password()
                user.save()
                sympathizer.user = user
                sympathizer.save()

            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))

            frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')
            link = f"{frontend_url}/set-password?uid={uid}&token={token}"

            logger.info(f"Password setup requested for cedula: {cedula[:4]}***")

            # Send email using EmailService
            if EmailService.send_password_setup(sympathizer, link):
                return Response({'message': 'Email enviado'})
            else:
                return Response(
                    {'error': 'Error al enviar el correo'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        except Sympathizer.DoesNotExist:
            logger.warning(f"Password setup requested for non-existent cedula: {cedula[:4]}***")
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)


class ForgotPasswordView(APIView):
    """Request password reset email for existing users."""

    @method_decorator(ratelimit(key='ip', rate='3/m', method='POST', block=True))
    def post(self, request):
        cedula = request.data.get('cedula')
        if not cedula:
            return Response({'error': 'La cedula es requerida'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sympathizer = Sympathizer.objects.select_related('user').get(cedula=cedula)

            if not sympathizer.email:
                return Response(
                    {'error': 'No hay correo electronico registrado para este usuario'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not sympathizer.user or not sympathizer.user.has_usable_password():
                return Response(
                    {'error': 'Esta cuenta no tiene una contrasena configurada. Use la opcion de configurar contrasena.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            user = sympathizer.user
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))

            frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')
            link = f"{frontend_url}/reset-password?uid={uid}&token={token}"

            logger.info(f"Password reset requested for cedula: {cedula[:4]}***")

            if EmailService.send_password_reset(sympathizer, link):
                return Response({'message': 'Email de recuperacion enviado'})
            else:
                return Response(
                    {'error': 'Error al enviar el correo'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        except Sympathizer.DoesNotExist:
            # Return success anyway to prevent enumeration attacks
            logger.info(f"Password reset requested for non-existent cedula")
            return Response({'message': 'Si el usuario existe, se enviara un correo de recuperacion'})


class SetPasswordView(APIView):
    """Set or reset password using token."""

    @method_decorator(ratelimit(key='ip', rate='10/m', method='POST', block=True))
    def post(self, request):
        uidb64 = request.data.get('uid')
        token = request.data.get('token')
        password = request.data.get('password')

        if not uidb64 or not token or not password:
            return Response({'error': 'Datos incompletos'}, status=status.HTTP_400_BAD_REQUEST)

        # Password validation
        if len(password) < 8:
            return Response(
                {'error': 'La contrasena debe tener al menos 8 caracteres'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            logger.warning(f"Invalid password set attempt with uid: {uidb64}")
            return Response({'error': 'Enlace invalido'}, status=status.HTTP_400_BAD_REQUEST)

        if default_token_generator.check_token(user, token):
            user.set_password(password)
            user.is_active = True
            user.save()

            # Update activated_at timestamp
            if hasattr(user, 'sympathizer'):
                user.sympathizer.activated_at = timezone.now()
                user.sympathizer.save()
                logger.info(f"Password set successfully for user: {user.username[:4]}***")

            return Response({'message': 'Contrasena configurada exitosamente'})
        else:
            logger.warning(f"Invalid or expired token for user: {user.username[:4]}***")
            return Response({'error': 'Token invalido o expirado'}, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    """User login endpoint."""

    @method_decorator(ratelimit(key='ip', rate='10/m', method='POST', block=True))
    def post(self, request):
        cedula = request.data.get('cedula')
        password = request.data.get('password')

        if not cedula or not password:
            return Response({'error': 'Cedula y contrasena son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=cedula, password=password)
        if user:
            try:
                # Check if sympathizer is suspended
                if hasattr(user, 'sympathizer') and user.sympathizer.is_suspended:
                    logger.warning(f"Login attempt by suspended user: {cedula[:4]}***")
                    return Response(
                        {'error': 'Tu cuenta esta suspendida. Contacta al administrador.'},
                        status=status.HTTP_403_FORBIDDEN
                    )

                token, _ = Token.objects.get_or_create(user=user)
                logger.info(f"Successful login for user: {cedula[:4]}***")
                return Response({
                    'token': token.key,
                    'user': {
                        'nombres': user.sympathizer.nombres,
                        'apellidos': user.sympathizer.apellidos,
                        'referral_code': user.sympathizer.referral_code
                    }
                })
            except Exception as e:
                logger.error(f"Error during login for {cedula[:4]}***: {str(e)}")
                return Response(
                    {'error': 'Error de inicio de sesion'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            logger.warning(f"Failed login attempt for cedula: {cedula[:4]}***")
            return Response({'error': 'Credenciales invalidas'}, status=status.HTTP_401_UNAUTHORIZED)


class DashboardView(APIView):
    """User dashboard with referral data."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            sympathizer = request.user.sympathizer

            # Build full referral tree with batched queries (all levels)
            children_map = defaultdict(list)
            queue = [sympathizer.id]
            visited = {sympathizer.id}

            while queue:
                batch_ids = queue
                queue = []
                children = Sympathizer.objects.filter(referrer_id__in=batch_ids).only(
                    'id', 'nombres', 'apellidos', 'cedula', 'email', 'phone', 'created_at', 'referrer_id'
                )

                for child in children:
                    if child.id in visited:
                        continue
                    visited.add(child.id)
                    children_map[child.referrer_id].append(child)
                    queue.append(child.id)

            # Keep newest first per parent
            for parent_id, child_list in children_map.items():
                child_list.sort(key=lambda item: item.created_at, reverse=True)

            def serialize_referral(ref):
                sub_children = children_map.get(ref.id, [])
                sub_referrals_data = [serialize_referral(sub) for sub in sub_children]
                return {
                    'id': ref.id,
                    'nombres': ref.nombres,
                    'apellidos': ref.apellidos,
                    'cedula': ref.cedula,
                    'email': ref.email,
                    'phone': ref.phone,
                    'created_at': ref.created_at,
                    'referrals_count': len(sub_children),
                    'sub_referrals': sub_referrals_data
                }

            referrals_data = [serialize_referral(ref) for ref in children_map.get(sympathizer.id, [])]

            return Response({
                'nombres': sympathizer.nombres,
                'apellidos': sympathizer.apellidos,
                'referral_code': sympathizer.referral_code,
                'referrals_count': len(referrals_data),
                'referrals': referrals_data
            })
        except Sympathizer.DoesNotExist:
            logger.error(f"Dashboard accessed by user without sympathizer: {request.user.username}")
            return Response(
                {'error': 'Perfil de simpatizante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in DashboardView: {str(e)}")
            return Response({'error': 'Error al cargar el dashboard'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LevelLabelView(APIView):
    """CRUD for per-user level labels."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            sympathizer = request.user.sympathizer
            labels = LevelLabel.objects.filter(owner=sympathizer).order_by('level')
            return Response({
                'level_labels': {str(label.level): label.name for label in labels}
            })
        except Sympathizer.DoesNotExist:
            logger.error(f"Level labels accessed by user without sympathizer: {request.user.username}")
            return Response(
                {'error': 'Perfil de simpatizante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in LevelLabelView.get: {str(e)}")
            return Response({'error': 'Error al cargar nombres de niveles'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request):
        try:
            sympathizer = request.user.sympathizer
            level_labels = request.data.get('level_labels')

            if not isinstance(level_labels, dict):
                return Response(
                    {'error': 'level_labels debe ser un objeto'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for level_key, name in level_labels.items():
                try:
                    level = int(level_key)
                except (TypeError, ValueError):
                    return Response(
                        {'error': f'Nivel invalido: {level_key}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if level < 1:
                    return Response(
                        {'error': 'El nivel debe ser >= 1'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                cleaned_name = (name or '').strip()
                if not cleaned_name:
                    LevelLabel.objects.filter(owner=sympathizer, level=level).delete()
                    continue

                LevelLabel.objects.update_or_create(
                    owner=sympathizer,
                    level=level,
                    defaults={'name': cleaned_name}
                )

            labels = LevelLabel.objects.filter(owner=sympathizer).order_by('level')
            return Response({
                'level_labels': {str(label.level): label.name for label in labels}
            })
        except Sympathizer.DoesNotExist:
            logger.error(f"Level labels updated by user without sympathizer: {request.user.username}")
            return Response(
                {'error': 'Perfil de simpatizante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in LevelLabelView.put: {str(e)}")
            return Response({'error': 'Error al guardar nombres de niveles'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class NetworkView(APIView):
    """Network visualization data with optimized tree layout."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            me = request.user.sympathizer

            # Data structures for layout
            nodes_dict = {}
            children_map = {}

            # Helper to create node dict
            def create_node_data(sympathizer, type_node, level):
                return {
                    'id': sympathizer.id,
                    'nombres': sympathizer.nombres,
                    'apellidos': sympathizer.apellidos,
                    'cedula': sympathizer.cedula,
                    'telefono': sympathizer.phone,
                    'email': sympathizer.email,
                    'referrals_count': 0,  # Will be updated later
                    'type': type_node,
                    'level': level,
                    'x': 0,
                    'y': 0
                }

            nodes_dict[me.id] = create_node_data(me, 'me', 0)
            children_map[me.id] = []

            # Add Sponsor if exists
            if me.referrer:
                sponsor = me.referrer
                nodes_dict[sponsor.id] = create_node_data(sponsor, 'sponsor', -1)
                children_map[sponsor.id] = [me.id]

            # Optimized BFS with prefetch
            visited = {me.id}
            queue = [(me, 0)]

            while queue:
                current, level = queue.pop(0)

                # Get children with prefetching
                direct_referrals = current.referrals.only(
                    'id', 'nombres', 'apellidos', 'cedula', 'phone', 'email'
                ).order_by('id')

                current_children_ids = []
                for ref in direct_referrals:
                    if ref.id not in visited:
                        visited.add(ref.id)
                        nodes_dict[ref.id] = create_node_data(ref, 'referral', level + 1)
                        children_map[ref.id] = []
                        current_children_ids.append(ref.id)
                        queue.append((ref, level + 1))

                children_map[current.id] = current_children_ids
                # Update referrals count
                nodes_dict[current.id]['referrals_count'] = len(current_children_ids)

            # Calculate Layout (Reingold-Tilford simplified)
            next_leaf_x = 0
            X_SPACING = 100

            def calculate_layout(node_id):
                nonlocal next_leaf_x

                children = children_map.get(node_id, [])

                if not children:
                    nodes_dict[node_id]['x'] = next_leaf_x
                    next_leaf_x += X_SPACING
                else:
                    child_xs = []
                    for child_id in children:
                        calculate_layout(child_id)
                        child_xs.append(nodes_dict[child_id]['x'])
                    nodes_dict[node_id]['x'] = sum(child_xs) / len(child_xs)

            calculate_layout(me.id)

            # If sponsor exists, place it above 'me'
            if me.referrer and me.referrer.id in nodes_dict:
                nodes_dict[me.referrer.id]['x'] = nodes_dict[me.id]['x']

            # Construct Response
            final_nodes = list(nodes_dict.values())
            final_links = []

            for parent_id, children_ids in children_map.items():
                for child_id in children_ids:
                    final_links.append({'source': parent_id, 'target': child_id})

            return Response({'nodes': final_nodes, 'links': final_links})

        except Sympathizer.DoesNotExist:
            logger.error(f"Network accessed by user without sympathizer: {request.user.username}")
            return Response(
                {'error': 'Perfil de simpatizante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in NetworkView: {str(e)}")
            return Response({'error': 'Error al cargar la red'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ImportTemplateView(APIView):
    """Download import template Excel file."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Importacion"

        # Headers with styling
        headers = [
            ('Nombres', True),
            ('Apellidos', True),
            ('Cedula', True),
            ('Telefono', True),
            ('Sexo', True),
            ('Email', False),
            ('Departamento', False),
            ('Municipio', False),
        ]

        header_fill_required = PatternFill(start_color="005258", end_color="005258", fill_type="solid")
        header_fill_optional = PatternFill(start_color="8FE7E5", end_color="8FE7E5", fill_type="solid")
        header_font_required = Font(bold=True, color="FFFFFF")
        header_font_optional = Font(bold=True, color="000000")

        for col_num, (header, required) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = f"{header}*" if required else f"{header} (opcional)"
            cell.fill = header_fill_required if required else header_fill_optional
            cell.font = header_font_required if required else header_font_optional
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Example row
        example_data = [
            'Juan',
            'Perez',
            '12345678',
            '3001234567',
            'M',
            'juan@ejemplo.com',
            'Antioquia',
            'Medellin'
        ]
        for col_num, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Instructions sheet
        ws_instructions = wb.create_sheet(title="Instrucciones")
        instructions = [
            "INSTRUCCIONES DE USO",
            "",
            "Campos obligatorios (marcados con *):",
            "- Nombres: Nombres de la persona",
            "- Apellidos: Apellidos de la persona",
            "- Cedula: Numero de cedula (solo numeros, debe ser unico)",
            "- Telefono: Numero de telefono (10 digitos)",
            "- Sexo: M (Masculino), F (Femenino), O (Otro)",
            "",
            "Campos opcionales:",
            "- Email: Correo electronico valido",
            "- Departamento: Nombre exacto del departamento (ej: Antioquia)",
            "- Municipio: Nombre exacto del municipio (requiere Departamento)",
            "",
            "NOTAS:",
            "- Los usuarios importados podran configurar su contrasena posteriormente",
            "- Cada usuario recibira un codigo de referido automaticamente",
            "- Las cedulas duplicadas seran rechazadas",
        ]
        for row_num, text in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_num, column=1, value=text)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
        ws_instructions.column_dimensions['A'].width = 70

        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importacion.xlsx"'

        return response


class ImportReferralsView(APIView):
    """Import referrals from Excel file."""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            current_user = request.user.sympathizer
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Perfil de simpatizante no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        # Check if user has link enabled
        if not current_user.link_enabled:
            return Response({'error': 'Tu enlace de referido esta deshabilitado'}, status=status.HTTP_403_FORBIDDEN)

        # Get file
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se proporciono archivo'}, status=status.HTTP_400_BAD_REQUEST)

        # Get parent user
        parent_id = request.data.get('parent_id')
        if parent_id:
            try:
                parent_id = int(parent_id)
                parent_user = Sympathizer.objects.get(pk=parent_id)

                # Validate parent is in current user's network
                if not self._is_in_network(current_user, parent_user):
                    return Response(
                        {'error': 'El usuario destino no pertenece a tu red'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except (ValueError, Sympathizer.DoesNotExist):
                return Response({'error': 'Usuario destino no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        else:
            parent_user = current_user

        # Load workbook
        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            return Response({'error': 'Error al leer el archivo Excel'}, status=status.HTTP_400_BAD_REQUEST)

        # Process rows
        errors = []
        created = []
        row_num = 0

        # Get existing cedulas for validation
        existing_cedulas = set(Sympathizer.objects.values_list('cedula', flat=True))

        # Cache departments and municipalities
        departments_cache = {d.name.upper(): d for d in Department.objects.all()}
        municipalities_cache = {}
        for m in Municipality.objects.select_related('department').all():
            key = (m.department.name.upper(), m.name.upper())
            municipalities_cache[key] = m

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1

            # Skip empty rows
            if not any(row):
                continue

            # Extract values
            nombres = str(row[0]).strip() if row[0] else ''
            apellidos = str(row[1]).strip() if row[1] else ''
            cedula = str(row[2]).strip() if row[2] else ''
            telefono = str(row[3]).strip() if row[3] else ''
            sexo = str(row[4]).strip().upper() if row[4] else ''
            email = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            departamento = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            municipio = str(row[7]).strip() if len(row) > 7 and row[7] else ''

            # Validate required fields
            if not nombres:
                errors.append({'row': row_num + 1, 'error': 'Campo Nombres es obligatorio'})
                continue
            if not apellidos:
                errors.append({'row': row_num + 1, 'error': 'Campo Apellidos es obligatorio'})
                continue
            if not cedula:
                errors.append({'row': row_num + 1, 'error': 'Campo Cedula es obligatorio'})
                continue
            if not telefono:
                errors.append({'row': row_num + 1, 'error': 'Campo Telefono es obligatorio'})
                continue
            if not sexo:
                errors.append({'row': row_num + 1, 'error': 'Campo Sexo es obligatorio'})
                continue

            # Clean cedula (only numbers)
            cedula = re.sub(r'\D', '', cedula)
            if not cedula:
                errors.append({'row': row_num + 1, 'error': 'Cedula debe contener numeros'})
                continue

            # Check duplicate cedula
            if cedula in existing_cedulas:
                errors.append({'row': row_num + 1, 'cedula': cedula, 'error': 'Cedula ya existe en el sistema'})
                continue

            # Validate phone (10 digits)
            telefono = re.sub(r'\D', '', telefono)
            if len(telefono) != 10:
                errors.append({'row': row_num + 1, 'cedula': cedula, 'error': 'Telefono debe tener 10 digitos'})
                continue

            # Validate sexo
            if sexo not in ['M', 'F', 'O']:
                errors.append({'row': row_num + 1, 'cedula': cedula, 'error': 'Sexo debe ser M, F u O'})
                continue

            # Validate email if provided
            if email and not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
                errors.append({'row': row_num + 1, 'cedula': cedula, 'error': 'Formato de email invalido'})
                continue

            # Handle department and municipality
            department_obj = None
            municipio_obj = None

            if departamento:
                department_obj = departments_cache.get(departamento.upper())
                if not department_obj:
                    errors.append({'row': row_num + 1, 'cedula': cedula, 'error': f'Departamento no encontrado: {departamento}'})
                    continue

                if municipio:
                    municipio_obj = municipalities_cache.get((departamento.upper(), municipio.upper()))
                    if not municipio_obj:
                        errors.append({'row': row_num + 1, 'cedula': cedula, 'error': f'Municipio no encontrado: {municipio}'})
                        continue

            # Create sympathizer
            try:
                new_sympathizer = Sympathizer.objects.create(
                    nombres=nombres,
                    apellidos=apellidos,
                    cedula=cedula,
                    phone=telefono,
                    sexo=sexo,
                    email=email if email else None,
                    department=department_obj,
                    municipio=municipio_obj,
                    referrer=parent_user,
                    link_enabled=True
                )
                existing_cedulas.add(cedula)
                created.append({
                    'id': new_sympathizer.id,
                    'cedula': cedula,
                    'nombres': nombres,
                    'apellidos': apellidos,
                    'referral_code': new_sympathizer.referral_code
                })
            except Exception as e:
                logger.error(f"Error creating sympathizer: {str(e)}")
                errors.append({'row': row_num + 1, 'cedula': cedula, 'error': 'Error al crear usuario'})

        logger.info(f"Import completed by {current_user.cedula[:4]}***: {len(created)} created, {len(errors)} errors")

        return Response({
            'success': True,
            'total_processed': row_num,
            'imported': len(created),
            'errors': errors,
            'created': created
        })

    def _is_in_network(self, current_user, target_user):
        """Check if target_user is in current_user's network (is a descendant)."""
        if current_user.id == target_user.id:
            return True

        # BFS to find if target is a descendant
        visited = {current_user.id}
        queue = [current_user]

        while queue:
            user = queue.pop(0)
            for referral in user.referrals.only('id'):
                if referral.id == target_user.id:
                    return True
                if referral.id not in visited:
                    visited.add(referral.id)
                    queue.append(referral)

        return False
