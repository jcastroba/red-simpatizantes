"""
Admin views for managing networks and users.
Includes export functionality and rate limiting.
"""
import logging
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from django_ratelimit.decorators import ratelimit
from django.utils.decorators import method_decorator

from .models import Sympathizer, Department, Municipality
from .serializers import SympathizerSerializer

logger = logging.getLogger(__name__)


class AdminLoginView(APIView):
    """Admin login endpoint."""
    permission_classes = [permissions.AllowAny]

    @method_decorator(ratelimit(key='ip', rate='5/m', method='POST', block=True))
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        if not username or not password:
            return Response({'error': 'Usuario y contrasena son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=username, password=password)

        if user and user.is_staff:
            token, created = Token.objects.get_or_create(user=user)
            logger.info(f"Admin login successful: {username}")
            return Response({
                'token': token.key,
                'user': {
                    'username': user.username,
                    'email': user.email,
                    'is_superuser': user.is_superuser
                }
            })

        logger.warning(f"Failed admin login attempt for: {username}")
        return Response({'error': 'Credenciales invalidas o no es administrador'}, status=status.HTTP_401_UNAUTHORIZED)


class AdminNetworkListView(APIView):
    """List and create root networks."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        # Networks are defined by root sympathizers (no referrer)
        # Optimized with annotation for referral count
        roots = Sympathizer.objects.filter(
            referrer__isnull=True
        ).annotate(
            direct_referrals_count=Count('referrals')
        ).order_by('-created_at')

        data = []
        for root in roots:
            data.append({
                'id': root.id,
                'name': f"{root.nombres} {root.apellidos}",
                'cedula': root.cedula,
                'phone': root.phone,
                'email': root.email,
                'referral_code': root.referral_code,
                'network_name': root.network_name or f"Red de {root.nombres} {root.apellidos}",
                'created_at': root.created_at,
                'link_enabled': root.link_enabled,
                'is_suspended': root.is_suspended,
                'direct_referrals': root.direct_referrals_count,
                'total_network_size': root.get_total_network_size()
            })

        return Response(data)

    def post(self, request):
        """Create a new network (Root Sympathizer)."""
        data = request.data

        # Validate required fields
        required = ['nombres', 'apellidos', 'cedula', 'phone', 'sexo']
        for field in required:
            if not data.get(field):
                return Response({'error': f'El campo {field} es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Check if cedula exists
            if Sympathizer.objects.filter(cedula=data['cedula']).exists():
                return Response({'error': 'La cedula ya esta registrada'}, status=status.HTTP_400_BAD_REQUEST)

            # Get email, converting empty string to None
            email = data.get('email')
            if email == '':
                email = None

            sympathizer = Sympathizer.objects.create(
                nombres=data['nombres'],
                apellidos=data['apellidos'],
                cedula=data['cedula'],
                phone=data['phone'],
                email=email,
                sexo=data['sexo'],
                department_id=data.get('department_id'),
                municipio_id=data.get('municipio_id'),
                network_name=data.get('network_name'),  # Nombre de la red/político
                referrer=None,
                link_enabled=True
            )

            logger.info(f"New network created by admin: {sympathizer.cedula[:4]}*** - Network: {sympathizer.network_name}")

            return Response({
                'message': 'Red creada exitosamente',
                'referral_code': sympathizer.referral_code,
                'id': sympathizer.id,
                'network_name': sympathizer.network_name or f"Red de {sympathizer.full_name}"
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Error creating network: {str(e)}")
            return Response({'error': 'Error al crear la red'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminUserListView(APIView):
    """List users with search and pagination."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        query = request.query_params.get('q', '')
        network_id = request.query_params.get('network_id')
        status_filter = request.query_params.get('status')  # active, suspended, pending

        # Optimized query with select_related
        users = Sympathizer.objects.select_related(
            'department', 'municipio', 'referrer', 'user'
        ).annotate(
            referrals_count=Count('referrals')
        ).order_by('-created_at')

        if query:
            users = users.filter(
                Q(nombres__icontains=query) |
                Q(apellidos__icontains=query) |
                Q(cedula__icontains=query) |
                Q(email__icontains=query)
            )

        if status_filter == 'active':
            users = users.filter(user__is_active=True, is_suspended=False)
        elif status_filter == 'suspended':
            users = users.filter(Q(is_suspended=True) | Q(user__is_active=False))
        elif status_filter == 'pending':
            users = users.filter(user__isnull=True)

        # Pagination
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start = (page - 1) * page_size
        end = start + page_size
        total = users.count()

        serializer = SympathizerSerializer(users[start:end], many=True)

        return Response({
            'results': serializer.data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'pages': (total + page_size - 1) // page_size
        })


class AdminUserDetailView(APIView):
    """Get, update, or delete a specific user."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, pk):
        try:
            user = Sympathizer.objects.select_related(
                'department', 'municipio', 'referrer', 'user'
            ).get(pk=pk)
            serializer = SympathizerSerializer(user)

            # Add extra data
            data = serializer.data
            data['total_network_size'] = user.get_total_network_size()

            return Response(data)
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    def patch(self, request, pk):
        """Partial update of user."""
        try:
            sympathizer = Sympathizer.objects.get(pk=pk)

            # Updateable fields
            updateable = ['nombres', 'apellidos', 'phone', 'email', 'sexo', 'link_enabled', 'is_suspended']
            for field in updateable:
                if field in request.data:
                    setattr(sympathizer, field, request.data[field])

            sympathizer.save()
            logger.info(f"User updated by admin: {sympathizer.cedula[:4]}***")

            return Response(SympathizerSerializer(sympathizer).data)
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        try:
            sympathizer = Sympathizer.objects.get(pk=pk)
            cedula = sympathizer.cedula

            # Store user reference before deletion
            auth_user = sympathizer.user

            # Delete sympathizer first
            sympathizer.delete()

            # Then delete associated auth user if exists
            if auth_user:
                auth_user.delete()

            logger.info(f"User deleted by admin: {cedula[:4]}***")
            return Response({'message': 'Usuario eliminado'})
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)


class AdminToggleLinkView(APIView):
    """Toggle link_enabled status for a user."""
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, pk):
        try:
            user = Sympathizer.objects.get(pk=pk)
            user.link_enabled = not user.link_enabled
            user.save()

            logger.info(f"Link toggled for user {user.cedula[:4]}***: link_enabled={user.link_enabled}")
            return Response({'link_enabled': user.link_enabled})
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)


class AdminToggleSuspensionView(APIView):
    """Toggle suspension status for a user."""
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, pk):
        try:
            sympathizer = Sympathizer.objects.get(pk=pk)

            # Toggle is_suspended on Sympathizer
            sympathizer.is_suspended = not sympathizer.is_suspended
            sympathizer.save()

            # Also toggle is_active on User if exists
            if sympathizer.user:
                sympathizer.user.is_active = not sympathizer.is_suspended
                sympathizer.user.save()

            logger.info(f"Suspension toggled for user {sympathizer.cedula[:4]}***: is_suspended={sympathizer.is_suspended}")
            return Response({
                'is_suspended': sympathizer.is_suspended,
                'is_active': not sympathizer.is_suspended
            })
        except Sympathizer.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)


class AdminExportUsersView(APIView):
    """Export users to Excel/CSV."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        export_format = request.query_params.get('format', 'xlsx')
        query = request.query_params.get('q', '')

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Get users
        users = Sympathizer.objects.select_related(
            'department', 'municipio', 'referrer', 'user'
        ).order_by('-created_at')

        if query:
            users = users.filter(
                Q(nombres__icontains=query) |
                Q(apellidos__icontains=query) |
                Q(cedula__icontains=query)
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Simpatizantes"

        # Headers
        headers = [
            'ID', 'Nombres', 'Apellidos', 'Cedula', 'Email', 'Telefono',
            'Sexo', 'Departamento', 'Municipio', 'Codigo Referido',
            'Referidor', 'Link Activo', 'Suspendido', 'Cuenta Activa',
            'Fecha Registro', 'Fecha Activacion'
        ]
        ws.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = cell.font.copy(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Data rows
        for user in users:
            ws.append([
                user.id,
                user.nombres,
                user.apellidos,
                user.cedula,
                user.email or '',
                user.phone,
                user.get_sexo_display(),
                user.department.name if user.department else '',
                user.municipio.name if user.municipio else '',
                user.referral_code,
                f"{user.referrer.nombres} {user.referrer.apellidos}" if user.referrer else 'Raiz',
                'Si' if user.link_enabled else 'No',
                'Si' if user.is_suspended else 'No',
                'Si' if (user.user and user.user.is_active) else 'No',
                user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else '',
                user.activated_at.strftime('%Y-%m-%d %H:%M') if user.activated_at else ''
            ])

        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"simpatizantes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        logger.info(f"Users exported by admin: {users.count()} records")
        return response
